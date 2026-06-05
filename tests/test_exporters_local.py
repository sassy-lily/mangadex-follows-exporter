from __future__ import annotations

import csv

import pytest
from openpyxl import load_workbook

from exporter.config import CsvExportConfig, XlsxExportConfig
from exporter.errors import ExportError
from exporter.exporters.csv import CsvExporter
from exporter.exporters.paths import render_path
from exporter.exporters.xlsx import XlsxExporter
from exporter.models import Dataset, MangaRecord, ReadProgress


def _dataset() -> Dataset:
    return Dataset(
        records=[
            MangaRecord(
                uuid="u1",
                status="reading",
                attributes={
                    "title": {"ja": "T1"},
                    "altTitles": [{"en": "Alt"}],
                    "links": {"raw": "jp", "engtl": "en"},
                },
                personal_rating=7,
                global_rating=8.0,
                read_progress=ReadProgress(highest_chapter=5.0, highest_volume=1.0),
            )
        ]
    )


def test_csv_writes_header_and_rows(tmp_path):
    path = tmp_path / "out.csv"
    exporter = CsvExporter(CsvExportConfig(path=str(path)))
    result = exporter.export(_dataset(), dry_run=False)

    assert result.success
    assert path.exists()
    rows = list(csv.reader(path.open()))
    assert rows[0][0] == "uuid"
    assert rows[1][0] == "u1"
    assert result.details["rows"] == 1


def test_csv_dry_run_writes_nothing(tmp_path):
    path = tmp_path / "out.csv"
    exporter = CsvExporter(CsvExportConfig(path=str(path)))
    result = exporter.export(_dataset(), dry_run=True)

    assert not path.exists()
    assert result.details["dry_run"] is True
    assert result.details["rows"] == 1


def test_csv_error_if_exists(tmp_path):
    path = tmp_path / "out.csv"
    path.write_text("existing")
    exporter = CsvExporter(
        CsvExportConfig(path=str(path), on_existing="error-if-exists")
    )
    with pytest.raises(ExportError, match="already exists"):
        exporter.export(_dataset(), dry_run=False)


def test_csv_custom_delimiter(tmp_path):
    path = tmp_path / "out.tsv"
    exporter = CsvExporter(CsvExportConfig(path=str(path), delimiter="\t"))
    exporter.export(_dataset(), dry_run=False)
    assert "\t" in path.read_text().splitlines()[0]


def test_xlsx_writes_sheet(tmp_path):
    path = tmp_path / "out.xlsx"
    exporter = XlsxExporter(XlsxExportConfig(path=str(path), sheet_name="manga"))
    result = exporter.export(_dataset(), dry_run=False)

    assert result.success
    wb = load_workbook(path)
    assert wb.sheetnames == ["manga"]
    sheet = wb["manga"]
    assert sheet.cell(row=1, column=1).value == "uuid"
    assert sheet.cell(row=2, column=1).value == "u1"


def test_xlsx_dry_run_writes_nothing(tmp_path):
    path = tmp_path / "out.xlsx"
    exporter = XlsxExporter(XlsxExportConfig(path=str(path)))
    exporter.export(_dataset(), dry_run=True)
    assert not path.exists()


def test_datetime_placeholder_rendered():
    from datetime import datetime

    rendered = render_path(
        "x-{datetime}.csv", now=datetime(2026, 6, 6, 13, 5, 9)
    )
    assert rendered.name == "x-2026-06-06T130509.csv"


def test_date_placeholder_rendered():
    from datetime import datetime

    rendered = render_path("x-{date}.csv", now=datetime(2026, 6, 6))
    assert rendered.name == "x-2026-06-06.csv"
